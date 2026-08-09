import unittest
from io import BytesIO

from openpyxl import load_workbook

from src.application.reporting import (
    EXCEL_HEADERS,
    REPORT_SECTION_LABELS,
    ClientReport,
    ReportSection,
    build_client_report,
    build_report_excel,
    render_report_html,
)
from src.storage.new_clients import (
    NewClient,
    ProcessingStatus,
    TelegramSearchAttempt,
)


def client(**changes) -> NewClient:
    values = {
        "spp_id": 1,
        "name": 'ООО "Тест"',
        "region": "Москва",
        "ogrn": "1127746271355",
        "inn": "7736641983",
        "kpp": "773601001",
        "is_entrepreneur": False,
        "registration_date": "2026-08-02",
        "liquidation_date": None,
        "director_last_name": "Иванов",
        "director_first_name": "Иван",
        "director_middle_name": "Иванович",
        "sbis_phones": ("+79990000001",),
        "telegram_phones": (),
        "sbis_emails": ("office@example.test",),
        "telegram_emails": (),
        "status": ProcessingStatus.QUEUED,
        "personalised_phones": ("+79990000002",),
        "personalised_emails": ("director@example.test",),
    }
    values.update(changes)
    return NewClient(**values)


class BuildClientReportTests(unittest.TestCase):
    def test_classifies_every_supported_section(self) -> None:
        clients = (
            client(
                spp_id=1,
                status=ProcessingStatus.PROCESSED,
                telegram_phones=("+79990000003",),
            ),
            client(spp_id=2, status=ProcessingStatus.SKIPPED),
            client(spp_id=3, status=ProcessingStatus.NEEDS_REVIEW),
            client(spp_id=4, status=ProcessingStatus.RETRY_REQUIRED),
            client(spp_id=5, status=ProcessingStatus.QUEUED),
        )

        report = build_client_report(clients)

        self.assertEqual(
            tuple(entry.section for entry in report.entries),
            (
                ReportSection.FOUND,
                ReportSection.SKIPPED,
                ReportSection.NEEDS_REVIEW,
                ReportSection.RETRY_REQUIRED,
                ReportSection.QUEUED,
            ),
        )

    def test_processed_without_telegram_contacts_requires_review(self) -> None:
        report = build_client_report(
            (client(status=ProcessingStatus.PROCESSED),)
        )

        self.assertEqual(report.entries[0].section, ReportSection.NEEDS_REVIEW)
        self.assertEqual(
            report.entries[0].error_code,
            "processed_without_telegram_contacts",
        )

    def test_adds_codes_from_latest_attempt(self) -> None:
        attempt = TelegramSearchAttempt(
            client_spp_id=1,
            attempt_number=2,
            stage="inn_query",
            result_code="temporary_error",
            error_code="telegram_timeout",
            created_at="2026-08-09 10:00:00",
        )

        report = build_client_report((client(),), {1: attempt})

        self.assertEqual(report.entries[0].result_code, "temporary_error")
        self.assertEqual(report.entries[0].error_code, "telegram_timeout")


class HtmlReportTests(unittest.TestCase):
    def test_queued_clients_are_shown_only_as_aggregate_count(self) -> None:
        report = build_client_report(
            (client(name="Скрытое имя очереди", status=ProcessingStatus.QUEUED),)
        )

        text = "\n".join(render_report_html(report))

        self.assertIn("Ещё не обработано — 1", text)
        self.assertNotIn("Скрытое имя очереди", text)

    def test_escapes_user_data_and_renders_sections(self) -> None:
        report = build_client_report(
            (
                client(
                    name="<Организация & партнёры>",
                    director_last_name="<Иванов>",
                    status=ProcessingStatus.PROCESSED,
                    telegram_phones=("+79990000003",),
                ),
            )
        )

        text = "\n".join(render_report_html(report))

        self.assertIn("&lt;Организация &amp; партнёры&gt;", text)
        self.assertIn("&lt;Иванов&gt;", text)
        self.assertNotIn("<Организация", text)
        for label in REPORT_SECTION_LABELS.values():
            self.assertIn(label, text)

    def test_splits_only_between_entries_and_observes_limit(self) -> None:
        report = build_client_report(
            tuple(
                client(
                    spp_id=index,
                    name=f"Организация-{index}-" + "А" * 80,
                    status=ProcessingStatus.PROCESSED,
                    telegram_phones=(f"+7999000{index:04d}",),
                )
                for index in range(1, 8)
            )
        )

        chunks = render_report_html(report, max_length=700)

        self.assertGreater(len(chunks), 1)
        self.assertTrue(all(len(chunk) <= 700 for chunk in chunks))
        combined = "\n".join(chunks)
        for index in range(1, 8):
            marker = f"Организация-{index}-"
            self.assertEqual(combined.count(marker), 1)

    def test_truncates_oversized_entry_without_splitting_it(self) -> None:
        report = build_client_report(
            (
                client(
                    name="Очень длинное название " * 500,
                    status=ProcessingStatus.PROCESSED,
                    telegram_phones=("+79990000003",),
                ),
            )
        )

        chunks = render_report_html(report)

        self.assertTrue(all(len(chunk) <= 4096 for chunk in chunks))
        self.assertEqual(sum("• <b>" in chunk for chunk in chunks), 1)
        self.assertIn("…", "".join(chunks))


class ExcelReportTests(unittest.TestCase):
    def test_builds_workbook_with_separate_contact_sources_and_codes(self) -> None:
        selected = client(
            status=ProcessingStatus.PROCESSED,
            telegram_phones=("+79990000003",),
            telegram_emails=("found@example.test",),
        )
        attempt = TelegramSearchAttempt(
            client_spp_id=1,
            attempt_number=1,
            stage="inn_report_parse",
            result_code="phone_found_by_inn",
            error_code=None,
            created_at="2026-08-09 10:00:00",
        )
        report = build_client_report((selected,), {1: attempt})

        content = build_report_excel(report)
        workbook = load_workbook(BytesIO(content), read_only=True)
        worksheet = workbook.active

        self.assertEqual(
            tuple(cell.value for cell in next(worksheet.iter_rows(max_row=1))),
            EXCEL_HEADERS,
        )
        row = tuple(cell.value for cell in next(worksheet.iter_rows(min_row=2)))
        self.assertEqual(row[0], REPORT_SECTION_LABELS[ReportSection.FOUND])
        self.assertEqual(row[1], 'ООО "Тест"')
        self.assertEqual(row[2], "Иванов Иван Иванович")
        self.assertEqual(row[5], "'+79990000001")
        self.assertEqual(row[6], "office@example.test")
        self.assertEqual(row[7], "'+79990000002")
        self.assertEqual(row[8], "director@example.test")
        self.assertEqual(row[9], "'+79990000003")
        self.assertEqual(row[10], "found@example.test")
        self.assertEqual(row[11], "phone_found_by_inn")
        self.assertIsNone(row[12])
        workbook.close()


if __name__ == "__main__":
    unittest.main()
