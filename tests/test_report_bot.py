"""Офлайн-тесты закрытого Telegram-бота отчетов."""

from __future__ import annotations

import unittest
from dataclasses import replace
from datetime import date
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

from aiogram.enums import ParseMode
from aiogram.types import BufferedInputFile
from openpyxl import load_workbook

from src.application.reporting import build_client_report
from src.cli.report_bot import DEFAULT_DATABASE_PATH, parse_arguments
from src.integrations.telegram.report_bot import (
    CALLBACK_ADMIN_CANCEL_PREFIX,
    CALLBACK_ADMIN_CONFIRM_PREFIX,
    CALLBACK_EXCEL_PREFIX,
    ReportBotService,
    admin_menu,
    create_report_router,
    load_report_bot_settings,
    menu_keyboard,
    parse_bootstrap_admin_ids,
    user_menu,
)
from src.integrations.telegram.report_sender import report_download_keyboard
from src.storage.new_clients import NewClient, ProcessingStatus


def make_message(
    user_id: int = 100,
    *,
    text: str = "",
    chat_type: str = "private",
) -> SimpleNamespace:
    """Создать минимальный объект сообщения с асинхронными ответами."""
    return SimpleNamespace(
        from_user=SimpleNamespace(id=user_id),
        chat=SimpleNamespace(type=chat_type),
        text=text,
        answer=AsyncMock(),
        answer_document=AsyncMock(),
        bot=SimpleNamespace(),
    )


def make_callback(
    message: SimpleNamespace,
    data: str,
    *,
    user_id: int | None = None,
) -> SimpleNamespace:
    """Создать callback с отдельно контролируемым Telegram ID."""
    return SimpleNamespace(
        from_user=SimpleNamespace(
            id=message.from_user.id if user_id is None else user_id
        ),
        message=message,
        data=data,
        answer=AsyncMock(),
    )


def make_client() -> NewClient:
    """Создать обезличенную карточку со всеми источниками контактов."""
    return NewClient(
        spp_id=17,
        name="ООО Тест & Партнеры",
        region="25",
        ogrn="1000000000000",
        inn="1000000000",
        kpp="100000001",
        is_entrepreneur=False,
        registration_date="2026-08-01T10:30:00",
        liquidation_date=None,
        director_last_name="Иванов",
        director_first_name="Иван",
        director_middle_name="Иванович",
        sbis_phones=("+70000000001",),
        telegram_phones=("+70000000002",),
        sbis_emails=("sbis@example.test",),
        telegram_emails=("tg@example.test",),
        status=ProcessingStatus.PROCESSED,
        personalised_phones=("+70000000003",),
        personalised_emails=("card@example.test",),
    )


def make_service(
    *,
    allowed: bool = True,
    admin: bool = False,
    clients: tuple[NewClient, ...] = (),
) -> tuple[ReportBotService, MagicMock, MagicMock]:
    """Создать сервис с синхронными mock-хранилищами."""
    access = MagicMock()
    access.is_user_allowed.return_value = allowed
    access.is_admin.return_value = admin
    access.list_subscribers.return_value = ()
    access.list_users.return_value = ()
    access.list_admins.return_value = (1,)
    access.subscribe.return_value = True
    access.unsubscribe.return_value = True
    access.add_user.return_value = True
    access.remove_user.return_value = True
    access.add_admin.return_value = True
    access.remove_admin.return_value = True
    access.latest_report_run.return_value = None
    access.latest_deliverable_report_run.return_value = None
    access.latest_pipeline_run.return_value = None
    access.list_integration_health.return_value = ()
    access.get_notification_state.return_value = None
    access.delivery_status_counts.return_value = {
        "pending": 0,
        "sent": 0,
        "failed": 0,
    }
    stored_runs: dict[int, SimpleNamespace] = {}
    stored_items: dict[int, tuple[object, ...]] = {}

    def create_report_run(**values: object) -> SimpleNamespace:
        report_id = len(stored_runs) + 1
        cohort_date = values["cohort_date"]
        run = SimpleNamespace(
            id=report_id,
            kind=values["kind"],
            cohort_date=(
                cohort_date.isoformat()
                if isinstance(cohort_date, date)
                else str(cohort_date)
            ),
            revision=values["revision"],
            created_at="2026-08-09T00:00:00+00:00",
        )
        stored_runs[report_id] = run
        stored_items[report_id] = tuple(values["items"])
        return run

    def latest_report_run(kind: str | None = None) -> SimpleNamespace | None:
        candidates = [
            run for run in stored_runs.values() if kind is None or run.kind == kind
        ]
        return max(candidates, key=lambda run: run.id, default=None)

    def find_report_run(**values: object) -> SimpleNamespace | None:
        for run in stored_runs.values():
            if (
                run.kind == values["kind"]
                and run.cohort_date
                == (
                    values["cohort_date"].isoformat()
                    if isinstance(values["cohort_date"], date)
                    else str(values["cohort_date"])
                )
                and run.revision == values["revision"]
            ):
                return run
        return None

    access.create_report_run.side_effect = create_report_run
    access.create_next_report_run.side_effect = lambda **values: create_report_run(
        **values,
        revision=(
            max(
                (
                    run.revision
                    for run in stored_runs.values()
                    if run.kind == values["kind"]
                    and run.cohort_date
                    == (
                        values["cohort_date"].isoformat()
                        if isinstance(values["cohort_date"], date)
                        else str(values["cohort_date"])
                    )
                ),
                default=0,
            )
            + 1
        ),
    )
    access.latest_report_run.side_effect = latest_report_run
    access.find_report_run.side_effect = find_report_run
    access.get_report_run.side_effect = stored_runs.get
    access.list_report_items.side_effect = lambda report_id: list(
        stored_items.get(report_id, ())
    )

    client_storage = MagicMock()
    client_storage.list_by_registration_date.return_value = list(clients)
    client_storage.latest_attempts_for_clients.return_value = {}
    service = ReportBotService(
        access,
        client_storage,
        bootstrap_admin_ids={1},
    )
    return service, access, client_storage


class SettingsTests(unittest.TestCase):
    def test_parses_unique_positive_bootstrap_admin_ids(self) -> None:
        self.assertEqual(
            parse_bootstrap_admin_ids(" 12,34,12 "),
            frozenset({12, 34}),
        )

    def test_rejects_empty_or_invalid_bootstrap_admin_ids(self) -> None:
        for value in ("", "0", "-1", "one"):
            with self.subTest(value=value), self.assertRaises(ValueError):
                parse_bootstrap_admin_ids(value)

    def test_loads_public_report_bot_settings(self) -> None:
        values = {
            "TELEGRAM_REPORT_BOT_TOKEN": "123456:TEST_TOKEN",
            "REPORT_BOT_BOOTSTRAP_ADMIN_IDS": "11,22",
        }
        with (
            patch("src.integrations.telegram.report_bot.loadEnvironment") as load,
            patch(
                "src.integrations.telegram.report_bot.requireSetting",
                side_effect=lambda name: values[name],
            ),
            patch.dict(
                "src.integrations.telegram.report_bot.os.environ",
                {
                    "REPORT_LOW_QUERY_THRESHOLD": "7",
                    "REPORT_RETENTION_DAYS": "120",
                },
            ),
        ):
            settings = load_report_bot_settings("state.sqlite3")

        load.assert_called_once_with()
        self.assertEqual(settings.token, "123456:TEST_TOKEN")
        self.assertEqual(settings.database_path, Path("state.sqlite3"))
        self.assertEqual(settings.bootstrap_admin_ids, frozenset({11, 22}))
        self.assertEqual(settings.low_query_threshold, 7)
        self.assertEqual(settings.report_retention_days, 120)

    def test_rejects_non_positive_low_query_threshold(self) -> None:
        values = {
            "TELEGRAM_REPORT_BOT_TOKEN": "123456:TEST_TOKEN",
            "REPORT_BOT_BOOTSTRAP_ADMIN_IDS": "11",
        }
        with (
            patch("src.integrations.telegram.report_bot.loadEnvironment"),
            patch(
                "src.integrations.telegram.report_bot.requireSetting",
                side_effect=lambda name: values[name],
            ),
            patch.dict(
                "src.integrations.telegram.report_bot.os.environ",
                {"REPORT_LOW_QUERY_THRESHOLD": "0"},
            ),
            self.assertRaises(ValueError),
        ):
            load_report_bot_settings("state.sqlite3")

    def test_rejects_negative_retention_days(self) -> None:
        values = {
            "TELEGRAM_REPORT_BOT_TOKEN": "123456:TEST_TOKEN",
            "REPORT_BOT_BOOTSTRAP_ADMIN_IDS": "11",
        }
        with (
            patch("src.integrations.telegram.report_bot.loadEnvironment"),
            patch(
                "src.integrations.telegram.report_bot.requireSetting",
                side_effect=lambda name: values[name],
            ),
            patch.dict(
                "src.integrations.telegram.report_bot.os.environ",
                {
                    "REPORT_LOW_QUERY_THRESHOLD": "10",
                    "REPORT_RETENTION_DAYS": "-1",
                },
            ),
            self.assertRaises(ValueError),
        ):
            load_report_bot_settings("state.sqlite3")

    def test_missing_retention_disables_cleanup(self) -> None:
        values = {
            "TELEGRAM_REPORT_BOT_TOKEN": "123456:TEST_TOKEN",
            "REPORT_BOT_BOOTSTRAP_ADMIN_IDS": "11",
        }
        with (
            patch("src.integrations.telegram.report_bot.loadEnvironment"),
            patch(
                "src.integrations.telegram.report_bot.requireSetting",
                side_effect=lambda name: values[name],
            ),
            patch.dict(
                "src.integrations.telegram.report_bot.os.environ",
                {},
                clear=True,
            ),
        ):
            settings = load_report_bot_settings("state.sqlite3")

        self.assertIsNone(settings.report_retention_days)

    def test_zero_retention_disables_cleanup(self) -> None:
        values = {
            "TELEGRAM_REPORT_BOT_TOKEN": "123456:TEST_TOKEN",
            "REPORT_BOT_BOOTSTRAP_ADMIN_IDS": "11",
        }
        with (
            patch("src.integrations.telegram.report_bot.loadEnvironment"),
            patch(
                "src.integrations.telegram.report_bot.requireSetting",
                side_effect=lambda name: values[name],
            ),
            patch.dict(
                "src.integrations.telegram.report_bot.os.environ",
                {"REPORT_RETENTION_DAYS": "0"},
            ),
        ):
            settings = load_report_bot_settings("state.sqlite3")

        self.assertIsNone(settings.report_retention_days)


class CliTests(unittest.TestCase):
    def test_cli_uses_project_database_by_default(self) -> None:
        arguments = parse_arguments([])
        self.assertEqual(arguments.database, DEFAULT_DATABASE_PATH)

    def test_cli_accepts_explicit_database(self) -> None:
        arguments = parse_arguments(["--database", "runtime/report.sqlite3"])
        self.assertEqual(arguments.database, Path("runtime/report.sqlite3"))


class KeyboardTests(unittest.TestCase):
    def test_persistent_keyboard_contains_menu_button(self) -> None:
        keyboard = menu_keyboard()
        self.assertTrue(keyboard.is_persistent)
        self.assertEqual(keyboard.keyboard[0][0].text, "Меню")

    def test_all_callback_data_fit_telegram_limit(self) -> None:
        keyboards = (
            user_menu(subscribed=False, is_admin=True),
            user_menu(subscribed=True, is_admin=False),
            admin_menu(),
            report_download_keyboard(7),
        )
        callbacks = [
            button.callback_data
            for keyboard in keyboards
            for row in keyboard.inline_keyboard
            for button in row
            if button.callback_data is not None
        ]
        self.assertTrue(callbacks)
        self.assertTrue(all(len(item.encode("utf-8")) <= 64 for item in callbacks))
        self.assertIn(f"{CALLBACK_EXCEL_PREFIX}7", callbacks)

    def test_router_is_created_without_network(self) -> None:
        service, _, _ = make_service()
        router = create_report_router(service)
        self.assertEqual(router.name, "report_bot")


class AccessHandlerTests(unittest.IsolatedAsyncioTestCase):
    async def test_unknown_private_user_receives_own_id(self) -> None:
        service, access, clients = make_service(allowed=False)
        message = make_message(987654)

        await service.start(message)

        sent_text = message.answer.await_args.args[0]
        self.assertIn("987654", sent_text)
        self.assertIn("администратору", sent_text)
        clients.list_by_registration_date.assert_not_called()
        access.is_user_allowed.assert_called_once_with(
            987654, bootstrap_admin_ids=frozenset({1})
        )

    async def test_id_is_available_without_whitelist(self) -> None:
        service, access, _ = make_service(allowed=False)
        message = make_message(345)

        await service.show_id(message)

        self.assertIn("345", message.answer.await_args.args[0])
        access.is_user_allowed.assert_not_called()

    async def test_group_message_cannot_use_bot(self) -> None:
        service, access, _ = make_service()
        message = make_message(chat_type="group")

        await service.start(message)

        self.assertIn("только в личном", message.answer.await_args.args[0])
        access.is_user_allowed.assert_not_called()

    async def test_callback_is_answered_before_authorization_and_snapshot(self) -> None:
        events: list[str] = []
        service, access, _ = make_service(clients=(make_client(),))
        message = make_message()
        callback = make_callback(
            message,
            f"{CALLBACK_EXCEL_PREFIX}7",
        )
        callback.answer.side_effect = lambda: events.append("answer")
        access.is_user_allowed.side_effect = lambda *args, **kwargs: events.append(
            "auth"
        ) or True
        access.get_report_run.side_effect = lambda report_id: events.append(
            "snapshot"
        )

        await service.excel_callback(callback)

        self.assertEqual(events, ["answer", "auth", "snapshot"])

    async def test_callback_rechecks_whitelist_and_denies_removed_user(self) -> None:
        service, _, clients = make_service(allowed=False)
        message = make_message(999)
        callback = make_callback(
            message,
            f"{CALLBACK_EXCEL_PREFIX}7",
        )

        await service.excel_callback(callback)

        callback.answer.assert_awaited_once_with()
        self.assertIn("999", message.answer.await_args.args[0])
        clients.list_by_registration_date.assert_not_called()
        service.access_storage.get_report_run.assert_not_called()
        message.answer_document.assert_not_awaited()

    async def test_inaccessible_callback_message_is_ignored_after_answer(self) -> None:
        service, access, _ = make_service()
        callback = make_callback(
            SimpleNamespace(chat=SimpleNamespace(type="private")),
            "status",
            user_id=100,
        )

        await service.status_callback(callback)

        callback.answer.assert_awaited_once_with()
        access.is_user_allowed.assert_not_called()


class ReportHandlerTests(unittest.IsolatedAsyncioTestCase):
    async def test_manual_report_reads_only_existing_sqlite_rows(self) -> None:
        service, access, clients = make_service(clients=(make_client(),))
        message = make_message(text="/report 2026-08-01")

        await service.report_command(message)

        clients.list_by_registration_date.assert_called_once_with(date(2026, 8, 1))
        attempts_call = clients.latest_attempts_for_clients.call_args.args[0]
        self.assertEqual(tuple(attempts_call), (17,))
        sent = "\n".join(call.args[0] for call in message.answer.await_args_list)
        self.assertIn("ООО Тест &amp; Партнеры", sent)
        self.assertIn("СБИС — телефоны", sent)
        self.assertIn("Telegram — телефоны", sent)
        last_options = message.answer.await_args.kwargs
        self.assertEqual(last_options["parse_mode"], ParseMode.HTML)
        button = last_options["reply_markup"].inline_keyboard[0][0]
        self.assertEqual(button.callback_data, "report:xlsx:id:1")
        access.create_next_report_run.assert_called_once()
        self.assertEqual(
            access.create_next_report_run.call_args.kwargs["kind"], "manual"
        )
        access.list_report_items.assert_called_once_with(1)

    async def test_report_without_date_does_not_query_database(self) -> None:
        service, _, clients = make_service()
        message = make_message(text="/report")

        await service.report_command(message)

        self.assertIn("YYYY-MM-DD", message.answer.await_args.args[0])
        clients.list_by_registration_date.assert_not_called()

    async def test_excel_uses_same_snapshot_after_client_changes(self) -> None:
        original = make_client()
        service, _, clients = make_service(clients=(original,))
        report_message = make_message(text="/report 2026-08-01")
        await service.report_command(report_message)

        clients.list_by_registration_date.return_value = [
            replace(original, name="ООО Измененное после HTML")
        ]
        clients.list_by_registration_date.reset_mock()
        download_message = make_message()
        callback = make_callback(download_message, "report:xlsx:id:1")

        await service.excel_callback(callback)

        callback.answer.assert_awaited_once_with()
        clients.list_by_registration_date.assert_not_called()
        document = download_message.answer_document.await_args.args[0]
        self.assertIsInstance(document, BufferedInputFile)
        self.assertEqual(document.filename, "report_1_2026-08-01.xlsx")
        self.assertTrue(document.data.startswith(b"PK"))
        workbook = load_workbook(BytesIO(document.data), read_only=True)
        values = tuple(
            cell
            for row in workbook.active.iter_rows(values_only=True)
            for cell in row
        )
        workbook.close()
        self.assertIn(original.name, values)
        self.assertNotIn("ООО Измененное после HTML", values)

    async def test_each_manual_request_uses_atomic_next_revision_api(self) -> None:
        service, access, _ = make_service(clients=(make_client(),))

        await service.report_command(make_message(text="/report 2026-08-01"))
        await service.report_command(make_message(text="/report 2026-08-01"))

        self.assertEqual(access.create_next_report_run.call_count, 2)
        access.find_report_run.assert_not_called()

    async def test_scheduled_excel_uses_saved_report_snapshot(self) -> None:
        service, access, clients = make_service()
        access.get_report_run.side_effect = None
        access.get_report_run.return_value = SimpleNamespace(
            id=7,
            cohort_date="2026-08-01",
        )
        access.list_report_items.side_effect = None
        access.list_report_items.return_value = [MagicMock()]
        message = make_message()
        callback = make_callback(message, "report:xlsx:id:7")
        snapshot = build_client_report((make_client(),))

        with patch(
            "src.integrations.telegram.report_bot.report_from_snapshot",
            return_value=snapshot,
        ) as restore:
            await service.excel_callback(callback)

        access.get_report_run.assert_called_once_with(7)
        access.list_report_items.assert_called_once_with(7)
        restore.assert_called_once()
        clients.list_by_registration_date.assert_not_called()
        document = message.answer_document.await_args.args[0]
        self.assertEqual(document.filename, "report_7_2026-08-01.xlsx")


class SubscriptionAndAdminTests(unittest.IsolatedAsyncioTestCase):
    async def test_subscribe_uses_bootstrap_admin_context(self) -> None:
        service, access, _ = make_service()
        message = make_message(123)

        await service.subscribe(message)

        access.subscribe.assert_called_once_with(
            123, bootstrap_admin_ids=frozenset({1})
        )
        self.assertIn("Подписка включена", message.answer.await_args.args[0])

    async def test_non_admin_cannot_execute_add_command(self) -> None:
        service, access, _ = make_service(allowed=True, admin=False)
        message = make_message(123, text="/add 456")
        state = SimpleNamespace(
            clear=AsyncMock(), set_state=AsyncMock(), update_data=AsyncMock()
        )

        await service.admin_command(message, state)

        callback_text = message.answer.await_args.args[0]
        self.assertIn("Ваш Telegram ID", callback_text)
        access.add_user.assert_not_called()
        state.clear.assert_awaited_once_with()

    async def test_admin_add_command_changes_whitelist(self) -> None:
        service, access, _ = make_service(admin=True)
        message = make_message(1, text="/add 456")
        state = SimpleNamespace(
            clear=AsyncMock(), set_state=AsyncMock(), update_data=AsyncMock()
        )

        await service.admin_command(message, state)

        access.add_user.assert_called_once_with(456, actor_user_id=1)
        self.assertIn("456", message.answer.await_args.args[0])

    async def test_admin_role_requires_one_time_bound_confirmation(self) -> None:
        service, access, _ = make_service(admin=True)
        message = make_message(1, text="/admin_add 456")
        state = SimpleNamespace(
            clear=AsyncMock(), set_state=AsyncMock(), update_data=AsyncMock()
        )

        await service.admin_command(message, state)

        access.add_admin.assert_not_called()
        keyboard = message.answer.await_args.kwargs["reply_markup"]
        confirm_data = keyboard.inline_keyboard[0][0].callback_data
        self.assertTrue(confirm_data.startswith(CALLBACK_ADMIN_CONFIRM_PREFIX))
        self.assertNotIn("456", confirm_data)

        foreign_callback = make_callback(message, confirm_data, user_id=2)
        await service.admin_confirmation_callback(foreign_callback)
        access.add_admin.assert_not_called()

        callback = make_callback(message, confirm_data, user_id=1)
        await service.admin_confirmation_callback(callback)
        access.add_admin.assert_called_once_with(
            456,
            actor_user_id=1,
            bootstrap_admin_ids=frozenset({1}),
        )

        await service.admin_confirmation_callback(callback)
        self.assertEqual(access.add_admin.call_count, 1)

    async def test_sensitive_admin_action_can_be_cancelled(self) -> None:
        service, access, _ = make_service(admin=True)
        message = make_message(1, text="/remove 456")
        state = SimpleNamespace(
            clear=AsyncMock(), set_state=AsyncMock(), update_data=AsyncMock()
        )

        await service.admin_command(message, state)

        keyboard = message.answer.await_args.kwargs["reply_markup"]
        cancel_data = keyboard.inline_keyboard[0][1].callback_data
        self.assertTrue(cancel_data.startswith(CALLBACK_ADMIN_CANCEL_PREFIX))
        callback = make_callback(message, cancel_data, user_id=1)
        await service.admin_confirmation_callback(callback)

        access.remove_user.assert_not_called()
        self.assertIn("отменено", message.answer.await_args.args[0])

    async def test_admin_inline_action_answers_then_enters_fsm(self) -> None:
        service, _, _ = make_service(admin=True)
        message = make_message(1)
        callback = make_callback(message, "admin:admin_add", user_id=1)
        state = SimpleNamespace(
            clear=AsyncMock(), set_state=AsyncMock(), update_data=AsyncMock()
        )

        await service.admin_action_callback(callback, state)

        callback.answer.assert_awaited_once_with()
        state.set_state.assert_awaited_once()
        state.update_data.assert_awaited_once_with(admin_operation="add_admin")
        self.assertIn("Telegram ID", message.answer.await_args.args[0])

    async def test_report_status_is_available_only_to_admin(self) -> None:
        service, access, _ = make_service(admin=False)
        message = make_message(123, text="/report_status")

        await service.report_status(message)

        self.assertIn("Ваш Telegram ID", message.answer.await_args.args[0])
        access.latest_deliverable_report_run.assert_not_called()

    async def test_report_status_shows_latest_delivery_counts(self) -> None:
        service, access, _ = make_service(admin=True)
        access.latest_deliverable_report_run.return_value = SimpleNamespace(
            id=7,
            kind="weekly",
            cohort_date="2026-08-01",
            revision=2,
            created_at="2026-08-08T10:00:00+00:00",
        )
        access.delivery_status_counts.return_value = {
            "pending": 3,
            "sent": 5,
            "failed": 1,
        }
        message = make_message(1, text="/report_status")

        await service.report_status(message)

        text = message.answer.await_args.args[0]
        self.assertIn("2026-08-01", text)
        self.assertIn("отправлено — 5", text)
        self.assertIn("ожидает — 3", text)
        self.assertIn("ошибок — 1", text)
        access.delivery_status_counts.assert_called_once_with(7)

    async def test_report_status_callback_answers_before_storage(self) -> None:
        events: list[str] = []
        service, access, _ = make_service(admin=True)
        message = make_message(1)
        callback = make_callback(message, "admin:status", user_id=1)
        callback.answer.side_effect = lambda: events.append("answer")
        access.is_admin.side_effect = lambda *args, **kwargs: events.append(
            "auth"
        ) or True
        access.latest_deliverable_report_run.side_effect = lambda: events.append(
            "storage"
        )

        await service.report_status_callback(callback)

        self.assertEqual(events, ["answer", "auth", "storage"])
        self.assertIn("еще не формировались", message.answer.await_args.args[0])

    async def test_pipeline_status_is_available_only_to_admin(self) -> None:
        service, access, _ = make_service(admin=False)
        message = make_message(123, text="/pipeline_status")

        await service.pipeline_status(message)

        self.assertIn("Ваш Telegram ID", message.answer.await_args.args[0])
        access.latest_pipeline_run.assert_not_called()

    async def test_pipeline_status_shows_safe_operational_aggregates(self) -> None:
        service, access, _ = make_service(admin=True)
        access.latest_pipeline_run.return_value = SimpleNamespace(
            id=8,
            target_date="2026-08-08",
            status="failed",
            started_at="2026-08-09T01:00:00+00:00",
            finished_at="2026-08-09T01:05:00+00:00",
            collected_cards=12,
            processing_status_counts={"processed": 7, "retry_required": 2},
            available_queries=4,
            error_stage="reports",
            error_code="TimeoutError",
        )
        message = make_message(1, text="/pipeline_status")

        await service.pipeline_status(message)

        text = message.answer.await_args.args[0]
        self.assertIn("завершен с ошибкой", text)
        self.assertIn("2026-08-08", text)
        self.assertIn("Обработан — 7", text)
        self.assertIn("Доступно запросов: 4", text)
        self.assertIn("Этап ошибки: reports", text)
        self.assertIn("Код ошибки: TimeoutError", text)
        self.assertNotIn("+7999", text)
        self.assertNotIn("example.test", text)

    async def test_pipeline_status_callback_answers_before_storage(self) -> None:
        events: list[str] = []
        service, access, _ = make_service(admin=True)
        message = make_message(1)
        callback = make_callback(message, "admin:pipeline_status", user_id=1)
        callback.answer.side_effect = lambda: events.append("answer")
        access.is_admin.side_effect = lambda *args, **kwargs: events.append(
            "auth"
        ) or True
        access.latest_pipeline_run.side_effect = lambda: events.append("storage")

        await service.pipeline_status_callback(callback)

        self.assertEqual(events, ["answer", "auth", "storage"])
        self.assertIn("еще не запускался", message.answer.await_args.args[0])

    async def test_health_status_shows_integrations_and_available_queries(self) -> None:
        service, access, _ = make_service(admin=True)
        access.list_integration_health.return_value = (
            SimpleNamespace(
                integration="sbis",
                status="healthy",
                checked_at="2026-08-09T01:00:00+00:00",
                last_ok_at="2026-08-09T01:00:00+00:00",
                error_code=None,
            ),
            SimpleNamespace(
                integration="telethon",
                status="unauthorized",
                checked_at="2026-08-09T01:01:00+00:00",
                last_ok_at=None,
                error_code="session_unauthorized",
            ),
        )
        access.latest_pipeline_run.return_value = SimpleNamespace(
            available_queries=17
        )
        message = make_message(1, text="/health")

        await service.health_status(message)

        text = message.answer.await_args.args[0]
        self.assertIn("СБИС cookie: работает", text)
        self.assertIn("Telegram-сессия: нужна повторная авторизация", text)
        self.assertIn("Бот отчетов: еще не проверялось", text)
        self.assertIn("Доступно запросов: 17", text)

    async def test_menu_button_rechecks_access(self) -> None:
        service, access, _ = make_service(allowed=True)
        message = make_message(100, text="Меню")

        await service.menu(message)

        self.assertIn("Главное меню", message.answer.await_args.args[0])
        access.is_user_allowed.assert_called_once()

    @patch(
        "src.integrations.telegram.report_bot.run_active_health_probes",
        new_callable=AsyncMock,
    )
    async def test_health_refresh_saves_current_balance_without_search(
        self, run_probes: AsyncMock
    ) -> None:
        service, access, _ = make_service(admin=True)
        run_probes.return_value = (
            (
                SimpleNamespace(
                    integration="sbis", status="healthy", error_code=None
                ),
                SimpleNamespace(
                    integration="telethon", status="healthy", error_code=None
                ),
                SimpleNamespace(
                    integration="report_bot", status="healthy", error_code=None
                ),
            ),
            34,
        )
        message = make_message(1, text="/health_refresh")

        await service.health_refresh(message)

        run_probes.assert_awaited_once_with(message.bot)
        self.assertEqual(access.record_integration_health.call_count, 3)
        access.set_notification_state.assert_called_once_with(
            "health_available_queries", "34"
        )

    async def test_test_report_requests_date_without_external_calls(self) -> None:
        service, _, clients = make_service(admin=True)
        message = make_message(1)
        callback = make_callback(message, "admin:test_report", user_id=1)
        state = SimpleNamespace(
            clear=AsyncMock(), set_state=AsyncMock(), update_data=AsyncMock()
        )

        await service.admin_test_report_callback(callback, state)

        callback.answer.assert_awaited_once_with()
        state.set_state.assert_awaited_once()
        self.assertIn("только уже сохраненные", message.answer.await_args.args[0])
        clients.list_by_registration_date.assert_not_called()


if __name__ == "__main__":
    unittest.main()
