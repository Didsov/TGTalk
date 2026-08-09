"""Построение безопасных HTML- и Excel-отчётов по новым клиентам."""

from __future__ import annotations

from collections.abc import Callable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from enum import StrEnum
from html import escape
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.storage.new_clients import (
    STATUS_LABELS,
    NewClient,
    NewClientStorage,
    ProcessingStatus,
    TelegramSearchAttempt,
)


TELEGRAM_MESSAGE_LIMIT = 4096


class ReportSection(StrEnum):
    """Раздел итогового отчёта."""

    FOUND = "found"
    SKIPPED = "skipped"
    NEEDS_REVIEW = "needs_review"
    RETRY_REQUIRED = "retry_required"
    QUEUED = "queued"


REPORT_SECTION_ORDER: tuple[ReportSection, ...] = (
    ReportSection.FOUND,
    ReportSection.SKIPPED,
    ReportSection.NEEDS_REVIEW,
    ReportSection.RETRY_REQUIRED,
    ReportSection.QUEUED,
)


REPORT_SECTION_LABELS: dict[ReportSection, str] = {
    ReportSection.FOUND: "Контакты найдены",
    ReportSection.SKIPPED: "Контакты не найдены",
    ReportSection.NEEDS_REVIEW: "Требуется ручная проверка",
    ReportSection.RETRY_REQUIRED: "Техническая ошибка, будет повтор",
    ReportSection.QUEUED: "Ещё не обработано",
}


EXCEL_HEADERS: tuple[str, ...] = (
    "Раздел",
    "Организация",
    "ФИО директора",
    "Дата регистрации",
    "Статус обработки",
    "Телефоны СБИС",
    "Почты СБИС",
    "Телефоны карточки СБИС",
    "Почты карточки СБИС",
    "Телефоны Telegram",
    "Почты Telegram",
    "Код результата",
    "Код ошибки",
)


@dataclass(frozen=True)
class ReportEntry:
    """Одна организация с уже определённым разделом отчёта."""

    client: NewClient
    section: ReportSection
    result_code: str | None
    error_code: str | None

    @property
    def director_name(self) -> str:
        """Вернуть ФИО директора без лишних пробелов."""
        parts = (
            self.client.director_last_name,
            self.client.director_first_name,
            self.client.director_middle_name,
        )
        return " ".join(part.strip() for part in parts if part and part.strip())


@dataclass(frozen=True)
class ClientReport:
    """Неизменяемый результат классификации выбранных клиентов."""

    entries: tuple[ReportEntry, ...]

    def entries_for(self, section: ReportSection) -> tuple[ReportEntry, ...]:
        """Вернуть записи одного раздела в исходном порядке."""
        return tuple(entry for entry in self.entries if entry.section is section)

    def count(self, section: ReportSection) -> int:
        """Вернуть количество записей раздела."""
        return sum(entry.section is section for entry in self.entries)


def load_client_report(
    storage: NewClientStorage,
    target_date: date | str,
) -> ClientReport:
    """Собрать отчет только из уже сохраненных SQL-данных за указанную дату."""
    clients = storage.list_by_registration_date(target_date)
    attempts = storage.latest_attempts_for_clients(
        client.spp_id for client in clients
    )
    return build_client_report(clients, attempts)


def build_client_report(
    clients: Sequence[NewClient],
    latest_attempts: Mapping[int, TelegramSearchAttempt] | None = None,
) -> ClientReport:
    """Классифицировать уже выбранных из БД клиентов для отчёта.

    Функция ничего не читает из внешних сервисов и не изменяет переданные
    объекты. ``latest_attempts`` используется только для отображения кода
    результата и ошибки последней Telegram-попытки.
    """
    attempts = latest_attempts or {}
    entries: list[ReportEntry] = []
    for client in clients:
        attempt = attempts.get(client.spp_id)
        section, invariant_error = _classify_client(client)
        entries.append(
            ReportEntry(
                client=client,
                section=section,
                result_code=attempt.result_code if attempt is not None else None,
                error_code=(
                    attempt.error_code
                    if attempt is not None and attempt.error_code
                    else invariant_error
                ),
            )
        )
    return ClientReport(entries=tuple(entries))


def render_report_html(
    report: ClientReport,
    *,
    title: str = "Отчёт по новым организациям",
    max_length: int = TELEGRAM_MESSAGE_LIMIT,
) -> tuple[str, ...]:
    """Отрисовать отчёт в Telegram HTML и разбить его между организациями.

    Данные экранируются до вставки в HTML. Одна организация никогда не
    переносится между сообщениями; если её текст слишком велик, значения
    сокращаются с многоточием, а полные данные остаются в Excel-версии.
    """
    if not 256 <= max_length <= TELEGRAM_MESSAGE_LIMIT:
        raise ValueError("max_length должен быть от 256 до 4096")

    summary_parts = tuple(
        f"{REPORT_SECTION_LABELS[section]}: {report.count(section)}"
        for section in REPORT_SECTION_ORDER
        if report.count(section) > 0
    )
    summary = ", ".join(summary_parts) if summary_parts else "Организаций: 0"
    blocks = [f"<b>{escape(title)}</b>\n{escape(summary)}"]

    for section in REPORT_SECTION_ORDER:
        entries = report.entries_for(section)
        heading = (
            f"<b>{escape(REPORT_SECTION_LABELS[section])} — "
            f"{len(entries)}</b>"
        )
        if not entries:
            continue
        if section is ReportSection.QUEUED:
            blocks.append(heading)
            continue
        first_entry = _render_entry_html(entries[0], max_length=max_length)
        combined = f"{heading}\n\n{first_entry}"
        if len(combined) <= max_length:
            blocks.append(combined)
        else:
            blocks.append(heading)
            blocks.append(first_entry)
        blocks.extend(
            _render_entry_html(entry, max_length=max_length)
            for entry in entries[1:]
        )

    return _pack_blocks(blocks, max_length=max_length)


def build_report_excel(
    report: ClientReport,
    *,
    sheet_title: str = "Новые организации",
) -> bytes:
    """Сформировать XLSX-файл в памяти, не создавая временных файлов."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_title[:31] or "Отчёт"
    worksheet.append(EXCEL_HEADERS)

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for section in REPORT_SECTION_ORDER:
        for entry in report.entries_for(section):
            client = entry.client
            worksheet.append(
                (
                    REPORT_SECTION_LABELS[entry.section],
                    _excel_text(client.name),
                    _excel_text(entry.director_name),
                    _excel_text(client.registration_date or ""),
                    STATUS_LABELS[client.status],
                    _excel_contacts(client.sbis_phones),
                    _excel_contacts(client.sbis_emails),
                    _excel_contacts(client.personalised_phones),
                    _excel_contacts(client.personalised_emails),
                    _excel_contacts(client.telegram_phones),
                    _excel_contacts(client.telegram_emails),
                    _excel_text(entry.result_code or ""),
                    _excel_text(entry.error_code or ""),
                )
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column, width in {
        "A": 34,
        "B": 40,
        "C": 35,
        "D": 18,
        "E": 30,
        "F": 26,
        "G": 30,
        "H": 30,
        "I": 30,
        "J": 26,
        "K": 30,
        "L": 30,
        "M": 30,
    }.items():
        worksheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _classify_client(client: NewClient) -> tuple[ReportSection, str | None]:
    if client.status is ProcessingStatus.PROCESSED:
        if client.telegram_phones or client.telegram_emails:
            return ReportSection.FOUND, None
        return ReportSection.NEEDS_REVIEW, "processed_without_telegram_contacts"
    if client.status is ProcessingStatus.SKIPPED:
        return ReportSection.SKIPPED, None
    if client.status is ProcessingStatus.NEEDS_REVIEW:
        return ReportSection.NEEDS_REVIEW, None
    if client.status is ProcessingStatus.RETRY_REQUIRED:
        return ReportSection.RETRY_REQUIRED, None
    return ReportSection.QUEUED, None


def _render_entry_html(entry: ReportEntry, *, max_length: int) -> str:
    phones = _merge_contacts(
        entry.client.telegram_phones,
        entry.client.personalised_phones,
        entry.client.sbis_phones,
    )
    emails = _merge_contacts(
        entry.client.telegram_emails,
        entry.client.personalised_emails,
        entry.client.sbis_emails,
    )
    values = (
        entry.client.name,
        _display_date(entry.client.registration_date),
        _first_contact_html(phones, formatter=_format_phone),
        _first_contact_html(emails),
    )

    def render(value_limit: int | None) -> str:
        name = _escaped_value(values[0], max_length=value_limit)
        registration_date = escape(values[1])
        return (
            f"<b>{name}</b>\n"
            f"Дата регистрации {registration_date}\n"
            f"Телефон: {values[2]}\n"
            f"Почта: {values[3]}"
        )

    full = render(None)
    if len(full) <= max_length:
        return full

    largest_value = len(values[0])
    low, high = 0, largest_value
    fitted: str | None = None
    while low <= high:
        middle = (low + high) // 2
        candidate = render(middle)
        if len(candidate) <= max_length:
            fitted = candidate
            low = middle + 1
        else:
            high = middle - 1
    if fitted is None:
        raise ValueError("max_length слишком мал для одного элемента отчёта")
    return fitted


def _merge_contacts(*sources: Sequence[str]) -> tuple[str, ...]:
    """Объединить источники в порядке приоритета без повторов."""
    result: list[str] = []
    seen: set[str] = set()
    for source in sources:
        for value in source:
            clean = value.strip()
            key = clean.casefold()
            if not clean or key in seen:
                continue
            seen.add(key)
            result.append(clean)
    return tuple(result)


def _first_contact_html(
    contacts: Sequence[str],
    *,
    formatter: Callable[[str], str] | None = None,
) -> str:
    if not contacts:
        return "—"
    value = formatter(contacts[0]) if formatter is not None else contacts[0]
    suffix = f" и еще {len(contacts) - 1}" if len(contacts) > 1 else ""
    return f"<code>{escape(value)}</code>{suffix}"


def _format_phone(value: str) -> str:
    digits = "".join(character for character in value if character.isdigit())
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        return (
            f"+7 ({digits[1:4]})-{digits[4:7]}-"
            f"{digits[7:9]}-{digits[9:11]}"
        )
    return value


def _display_date(value: str | None) -> str:
    if not value:
        return "—"
    try:
        parsed = date.fromisoformat(value[:10])
    except ValueError:
        return value
    return parsed.strftime("%d.%m.%Y")


def _pack_blocks(blocks: Sequence[str], *, max_length: int) -> tuple[str, ...]:
    chunks: list[str] = []
    current = ""
    for block in blocks:
        if len(block) > max_length:
            raise ValueError("Элемент отчёта превышает max_length")
        candidate = block if not current else f"{current}\n\n{block}"
        if len(candidate) <= max_length:
            current = candidate
            continue
        chunks.append(current)
        current = block
    if current:
        chunks.append(current)
    return tuple(chunks)


def _contacts_text(contacts: Sequence[str]) -> str:
    return ", ".join(contacts) if contacts else "—"


def _escaped_value(value: str, *, max_length: int | None) -> str:
    if max_length is None:
        return escape(value)
    if max_length == 0:
        return "…"
    if len(value) > max_length:
        value = f"{value[:max(0, max_length - 1)]}…"
    return escape(value)


def _excel_contacts(contacts: Sequence[str]) -> str:
    return _excel_text("\n".join(contacts))


def _excel_text(value: str) -> str:
    """Защитить текстовую ячейку от интерпретации как Excel-формулы."""
    if value and value[0] in ("=", "+", "-", "@"):
        return f"'{value}"
    return value
